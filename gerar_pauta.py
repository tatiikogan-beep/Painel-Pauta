#!/usr/bin/env python3
"""
Módulo de geração da Reunião de Pauta.
Função principal: gerar_pauta(src_new_bytes, src_old_bytes) -> (output_bytes, resumo, divergencias)
"""
import io, datetime, unicodedata
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule as OXLRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
import openpyxl

# ── Paleta ────────────────────────────────────────────────────────────────────
NAVY='1F3864'; MED_BLUE='2E75B6'; SISTEMA_BG='D9E1F2'
ADVOG_BG='E2EFDA'; CTRL_BG='FCE4D6'
ORANGE_ROW='FFF2CC'; GREEN_ROW='E8F5E9'; GRAY_ALT='F5F5F5'
WHITE='FFFFFF'; RED_DARK='C00000'; BORDER_C='BDBDBD'; LIGHT_C='E0E0E0'
S2_DARK='6B7280'; S2_MED='9CA3AF'; S2_LIGHT='F3F4F6'; S2_ACCENT='E5E7EB'
DIV_RED='FFC7CE'; DIV_DARK='9C0006'

# ── Seriais de data ───────────────────────────────────────────────────────────
_EPOCH = datetime.date(1899,12,30)
def _serial(d): return (d - _EPOCH).days

def s1_cond(extra="", s1=None, s2=None):
    return f'GERAL!$A:$A,">="&{s1},GERAL!$A:$A,"<"&{s2}{extra}'
def s2_cond(extra="", s2=None, s3=None):
    return f'GERAL!$A:$A,">="&{s2},GERAL!$A:$A,"<"&{s3}{extra}'

def tb(c=BORDER_C):
    s=Side(style='thin',color=c); return Border(left=s,right=s,top=s,bottom=s)
def lb():
    s=Side(style='thin',color=LIGHT_C); return Border(left=s,right=s,top=s,bottom=s)

def hcell(ws,row,col,val,fg=WHITE,bg=NAVY,bold=True,size=9,align='center',wrap=True):
    c=ws.cell(row,col,val)
    c.font=Font(name='Arial',bold=bold,color=fg,size=size)
    c.fill=PatternFill('solid',start_color=bg)
    c.alignment=Alignment(horizontal=align,vertical='center',wrap_text=wrap)
    c.border=tb(); return c

# ── Mapeamento coordenadores ──────────────────────────────────────────────────
COORDS={
    'GABRIEL GIORGIO CICCHELERO':{'dark':'1F4E79','light':'DDEEFF',
        'adv':['ALYSSON NARBAL DE OLIVEIRA SOMBRA','ARMANDO HÉLIO ALMEIDA MONTEIRO DE MORAES',
               'IRENE FLÁVIA SERENÁRIO','JAMILE BARRETO','JULIANA DE OLIVEIRA ROCHA',
               'RAFAEL CAVALCANTE BARBOSA']},
    'HELANZIA DE ARAUJO XAVIER WICHMANN':{'dark':'7B3F00','light':'FFE5CC',
        'adv':['ARTUR SARAIVA DE ANDRADE','GUSTAVO LOPES ALENCAR FILHO','KELIANE DE OLIVEIRA',
               'NATALIA PAIVA DE PAULA','ROBERTA RAYANNE VASCONCELOS BOTO',
               'VICTOR EMANOEL FRADIQUE ACCIOLY FONTENELE','WELLINGTON PEREIRA DA ROCHA FILHO']},
    'JENIFFER ROSA BARBOSA DE SALES':{'dark':'4A235A','light':'F5D6FF',
        'adv':['PAULO MARCIO SOARES DE CARVALHO FILHO']},
    'JULIANA MIRELLA ALVES RODRIGUES':{'dark':'145A32','light':'D5F5E3',
        'adv':['THALLYS ANDERSON FERREIRA DE LIMA']},
    'LUCIANE MODERNEL MENDES':{'dark':'1B4F72','light':'D6EAF8',
        'adv':['ANTONIO EDUARDO GOES AGUIAR FILHO','EDUARDO BLASQUES MARTINE',
               'LAYLA EVELYN NASCIMENTO PINHEIRO','MATHEUS CAVALCANTI DE ARAUJO',
               'SANE BORGES BORGOMONI']},
    'MARCELLE LEITE RENTROIA':{'dark':'6E2C00','light':'FDEBD0','adv':['MARIANA MOTA FROTA']},
    'NAYANDERSON LUAN MELLO PINHEIRO':{'dark':'0B5345','light':'D1F2EB',
        'adv':['ANDRE VIANA GARRIDO','EMERSON TRAVASSOS TORQUATO','YURI GONDIM DE AMORIM']},
    'RONALD FEITOSA AGUIAR FILHO':{'dark':'512E5F','light':'F8C6FF',
        'adv':['ALEXIA ALENCAR CAPIBARIBE']},
    'SUZANA MARIA CAMPOS MARANHÃO DE LIMA':{'dark':'922B21','light':'FADBD8',
        'adv':['DANIEL BARROS DE OLIVEIRA','EVILANY GABRIELA BRAGA PONTES',
               'FRANCOISE CATHERINE SOUZA ALVES','GIOVANNA CAMPOS PEREIRA',
               'GIOVANNA CESAR FERREIRA','LETICIA OLIVEIRA DA SILVA','TATIANE CARMO SANTA ROSA']},
    'YURI ALVES BARROS DOS SANTOS':{'dark':'154360','light':'D6EAF8',
        'adv':['LUIZ GUILHERME GONCALVES GIRÃO']},
    'CONTROLADORIA JURÍDICA':{'dark':'7D6608','light':'FDFFD6',
        'adv':['CONTROLADORIA JURÍDICA','SUPORTE','AJ - CONTROLADORIA JURÍDICA']},
}

adv2coord={}
for coord,info in COORDS.items():
    adv2coord[coord]=coord
    for adv in info['adv']: adv2coord[adv]=coord

ALL_KNOWN_ADVS={a for d in COORDS.values() for a in d['adv']} | set(COORDS.keys())

# ── Normalização ──────────────────────────────────────────────────────────────
def _plain(s):
    return ''.join(c for c in unicodedata.normalize('NFD',s.upper())
                   if unicodedata.category(c)!='Mn')

def normalize_adv(val):
    if not val: return val
    v=str(val).strip()
    if 'CONTROLADORIA' in _plain(v): return 'CONTROLADORIA JURÍDICA'
    if 'KELIANE' in v.upper() and '/' in v: return 'KELIANE DE OLIVEIRA'
    if 'CAVALCANTE' in v.upper() and 'BARSOSA' in v.upper(): return 'RAFAEL CAVALCANTE BARBOSA'
    return v

def get_coord(resp_pasta):
    if pd.isna(resp_pasta) or not resp_pasta: return ''
    v=normalize_adv(str(resp_pasta).strip())
    if v=='ALEXIA ALENCAR CAPIBARIBE': return 'RONALD FEITOSA AGUIAR FILHO'
    return adv2coord.get(v, v)

# ── Detecção de divergências ──────────────────────────────────────────────────
def detectar_divergencias(df_new, df_old_ws=None, preserved=None):
    """Retorna lista de dicts com divergências encontradas."""
    divs=[]

    # 1. Advogados não reconhecidos em Responsável pela Pasta
    resp_vals=df_new['Responsável pela Pasta'].dropna().unique()
    for v in resp_vals:
        norm=normalize_adv(str(v).strip())
        if norm not in ALL_KNOWN_ADVS and norm not in adv2coord:
            cnt=int((df_new['Responsável pela Pasta'].astype(str).str.strip()==str(v).strip()).sum())
            divs.append({
                'tipo':'NOME NÃO RECONHECIDO',
                'campo':'Responsável pela Pasta',
                'valor':str(v),
                'valor_sugerido': normalize_adv(str(v)),
                'ocorrencias':cnt,
                'gravidade':'ALTA',
                'descricao':f'"{v}" não está na lista de advogados/coordenadores cadastrados'
            })

    # 2. Processos sem coordenador mapeável
    sem_coord=df_new[df_new['Responsável pela Pasta'].isna() |
                      (df_new['Responsável pela Pasta'].astype(str).str.strip()=='')]
    if len(sem_coord)>0:
        divs.append({
            'tipo':'SEM RESPONSÁVEL',
            'campo':'Responsável pela Pasta',
            'valor':'(vazio)',
            'valor_sugerido':'',
            'ocorrencias':len(sem_coord),
            'gravidade':'MÉDIA',
            'descricao':f'{len(sem_coord)} processo(s) sem Responsável pela Pasta'
        })

    # 3. Verificar variações ortográficas nos dados preservados
    if preserved:
        for key,rec in preserved.items():
            adv=rec.get('adv','')
            if adv and str(adv).strip():
                norm=normalize_adv(str(adv))
                if norm!=str(adv).strip() and norm:
                    divs.append({
                        'tipo':'GRAFIA CORRIGIDA',
                        'campo':'Advogado Responsável pela Audiência',
                        'valor':str(adv).strip(),
                        'valor_sugerido':norm,
                        'ocorrencias':1,
                        'gravidade':'BAIXA',
                        'descricao':f'Grafia "{adv}" corrigida para "{norm}"'
                    })

    return divs

# ── Preservação ───────────────────────────────────────────────────────────────
def carregar_preservados(src_old_bytes):
    preserved={}; preserved_2={}; preserved_1={}
    if not src_old_bytes: return preserved, preserved_2, preserved_1
    try:
        _wb=openpyxl.load_workbook(io.BytesIO(src_old_bytes),data_only=True)

        sem_idx={}; sem_contrat={}; sem_preposto={}
        if 'SEMANA' in _wb.sheetnames:
            ws=_wb['SEMANA']
            hdrs={ws.cell(2,c).value:c for c in range(1,ws.max_column+1)}
            sc_cnj=hdrs.get('Número de CNJ',4); sc_cli=hdrs.get('Cliente Processo',10)
            sc_g=hdrs.get('Dados dos Correspondentes',7)
            sc_f=hdrs.get('Contratação',6); sc_h=hdrs.get('Preposto',8)
            for r in range(3,ws.max_row+1):
                cnj=str(ws.cell(r,sc_cnj).value or '').strip()
                cli=str(ws.cell(r,sc_cli).value or '').strip()
                if not cnj: continue
                v=ws.cell(r,sc_g).value
                if v and str(v).strip() and v!=0: sem_idx[(cnj,cli)]=v
                v2=ws.cell(r,sc_f).value
                if v2 and str(v2).strip(): sem_contrat[(cnj,cli)]=v2
                v3=ws.cell(r,sc_h).value
                if v3 and str(v3).strip(): sem_preposto[(cnj,cli)]=v3

        if 'GERAL' in _wb.sheetnames:
            ws=_wb['GERAL']
            hdr_row=1 if any('CNJ' in str(ws.cell(1,c).value or '') for c in range(1,ws.max_column+1)) else 2
            dat_row=hdr_row+1
            hdrs={ws.cell(hdr_row,c).value:c for c in range(1,ws.max_column+1)}
            c_cnj=hdrs.get('Número de CNJ',4); c_adv=hdrs.get('Advogado Responsável pela Audiência',7)
            c_dados=hdrs.get('Dados dos Correspondentes',8); c_acomp=hdrs.get('Acompanhamento',9)
            c_obs=hdrs.get('Observações',10); c_cli=hdrs.get('Cliente Processo',11)
            c_data=hdrs.get('Data/hora de início',1); c_desc=hdrs.get('Descrição',3)
            for r in range(dat_row,ws.max_row+1):
                cnj=ws.cell(r,c_cnj).value
                if cnj:
                    cnj=str(cnj).strip(); cli=str(ws.cell(r,c_cli).value or '').strip()
                    data=str(ws.cell(r,c_data).value or '').strip()[:10]
                    desc=str(ws.cell(r,c_desc).value or '').strip()[:60]
                    adv=normalize_adv(ws.cell(r,c_adv).value)
                    raw=sem_idx.get((cnj,cli)) or ws.cell(r,c_dados).value
                    dados=raw if(raw and str(raw).strip() and raw!=0) else None
                    acomp=ws.cell(r,c_acomp).value; obs=ws.cell(r,c_obs).value
                    contrat=sem_contrat.get((cnj,cli)); preposto=sem_preposto.get((cnj,cli))
                    if any(v for v in [adv,dados,acomp,obs,contrat,preposto]):
                        rec={'adv':adv,'dados':dados,'acomp':acomp,'obs':obs,
                             'contrat':contrat,'preposto':preposto}
                        preserved[(cnj,cli,data,desc)]=rec
                        if (cnj,cli) not in preserved_2: preserved_2[(cnj,cli)]=rec
                        if cnj not in preserved_1: preserved_1[cnj]=rec
    except Exception as e:
        pass
    return preserved, preserved_2, preserved_1

# ── Geração principal ─────────────────────────────────────────────────────────
def gerar_pauta(src_new_bytes: bytes, src_old_bytes: bytes=None):
    """
    Gera a planilha Reunião de Pauta.
    Retorna: (output_bytes, resumo_dict, divergencias_list)
    """
    # Carregar fonte nova
    df_src=pd.read_excel(io.BytesIO(src_new_bytes),header=0)
    df_src.columns=[c.strip() for c in df_src.columns]
    df_src['Data/hora de início']=pd.to_datetime(df_src['Data/hora de início'],errors='coerce')
    df_src['Coordenador']=df_src['Responsável pela Pasta'].apply(get_coord)

    df_g=df_src[df_src['Status']=='Pendente'].copy()
    df_g=df_g.drop_duplicates(subset=['Número de CNJ','Cliente Processo'],keep='first').reset_index(drop=True)
    df_c=df_src[df_src['Status']=='Cancelado'].copy().reset_index(drop=True)

    # Semana
    today=datetime.date.today()
    days=(7-today.weekday())%7
    if days==0: days=7
    s1_start=today+datetime.timedelta(days=days)
    _Q='"'  # usado nas f-strings de fórmula
    _S1=_serial(s1_start); _S2=_S1+7; _S3=_S1+14

    # Preservação
    preserved, preserved_2, preserved_1=carregar_preservados(src_old_bytes)

    # Normalizar _active_advs
    _active_advs=set(normalize_adv(str(v).strip()) for v in df_g['Responsável pela Pasta'].dropna())

    # Detectar divergências
    divs=detectar_divergencias(df_g, preserved=preserved)

    # Construir planilha
    wb=Workbook()

    # CONFIG (oculta)
    ws_cfg=wb.active; ws_cfg.title='CONFIG'; ws_cfg.sheet_state='hidden'
    all_advs=sorted(ALL_KNOWN_ADVS|{'AJ - CONTROLADORIA JURÍDICA','CONTROLADORIA JURÍDICA','SUPORTE'})
    ws_cfg['A1']='ADVOGADOS'
    for i,a in enumerate(all_advs,2): ws_cfg.cell(i,1,a)
    ws_cfg['B1']='PRESENCIAL'; ws_cfg['B2']='VIRTUAL'
    ws_cfg.cell(3,2).value='HÍBRIDA'; ws_cfg['B4']='Sim'
    ws_cfg.cell(5,2).value='Não'
    ws_cfg['B6']='*CONTROLADORIA*'
    ws_cfg.cell(7,2).value='Tributária'; ws_cfg.cell(8,2).value='Cível'
    _P='CONFIG!$B$1'; _V='CONFIG!$B$2'; _H='CONFIG!$B$3'
    _SIM='CONFIG!$B$4'; _NAO='CONFIG!$B$5'; _CJ='CONFIG!$B$6'
    _TRIB='CONFIG!$B$7'; _CIVEL='CONFIG!$B$8'

    def sc1(extra=""): return s1_cond(extra,_S1,_S2)
    def sc2(extra=""): return s2_cond(extra,_S2,_S3)

    # ── ABA GERAL ──────────────────────────────────────────────────────────────
    ws_g=wb.create_sheet('GERAL',1); ws_g.sheet_properties.tabColor=MED_BLUE
    COLS_G=[
        ('Data/hora de início',18,'S'),('Natureza',14,'S'),('Descrição',46,'S'),
        ('Número de CNJ',24,'S'),('Tipo / Subtipo',22,'S'),('Coordenador',32,'S'),
        ('Advogado Responsável pela Audiência',35,'A'),('Dados dos Correspondentes',38,'F'),
        ('Acompanhamento',16,'A'),('Observações',32,'A'),('Cliente Processo',38,'S'),
        ('Contrário principal',30,'S'),('Modalidade',14,'A'),('Local',44,'A'),
        ('Classificação do Processo',22,'S'),('Advogado Responsável',30,'S'),
        ('Ação',25,'S'),('Órgão',42,'S'),('Outros envolvidos',25,'S'),
        ('Cidade',18,'S'),('UF',8,'S'),('Status',14,'S'),('_CANC',1,'S'),
    ]
    CSTY={'S':(SISTEMA_BG,'1F3864'),'A':(ADVOG_BG,'375623'),'F':(CTRL_BG,'843C0C')}
    WRAP_G={'Descrição','Local','Observações'}
    SRC_MAP={
        'Data/hora de início':'Data/hora de início','Natureza':'Natureza',
        'Descrição':'Descrição','Número de CNJ':'Número de CNJ',
        'Tipo / Subtipo':'Tipo / Subtipo','Coordenador':'Coordenador',
        'Cliente Processo':'Cliente Processo','Contrário principal':'Contrário principal',
        'Modalidade':'Modalidade','Local':'Local',
        'Classificação do Processo':'Classificação do Processo',
        'Advogado Responsável':'Responsável pela Pasta',
        'Ação':'Ação','Órgão':'Órgão','Outros envolvidos':'Outros envolvidos',
        'Cidade':'Cidade','UF':'UF','Status':'Status',
    }
    ws_g.row_dimensions[1].height=14; ws_g.row_dimensions[2].height=38
    for ci,(nome,larg,tipo) in enumerate(COLS_G,1):
        bg,fg=CSTY[tipo]
        label=nome if not nome.startswith('_') else ''
        c=ws_g.cell(2,ci,label)
        c.font=Font(name='Arial',bold=True,color=fg,size=9)
        c.fill=PatternFill('solid',start_color=bg)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=tb()
        ws_g.column_dimensions[get_column_letter(ci)].width=larg

    _resolved={}
    # Marcar linhas com divergência para highlight no GERAL
    div_cnjs={d['valor'] for d in divs if d['tipo']=='NOME NÃO RECONHECIDO'}

    for ri,row in df_g.iterrows():
        er=ri+3
        cnj=str(row.get('Número de CNJ','')).strip()
        cli=str(row.get('Cliente Processo','')).strip()
        data=str(row.get('Data/hora de início','')).strip()[:10]
        desc=str(row.get('Descrição','')).strip()[:60]
        resp=str(row.get('Responsável pela Pasta','')).strip()
        prsv=(preserved.get((cnj,cli,data,desc)) or preserved_2.get((cnj,cli)) or preserved_1.get(cnj) or {})
        _resolved[ri]={'adv':prsv.get('adv'),'dados':prsv.get('dados'),
                       'acomp':prsv.get('acomp'),'obs':prsv.get('obs')}
        is_div=resp in div_cnjs  # linha com divergência

        for ci,(nome,_,tipo) in enumerate(COLS_G,1):
            val=None
            if nome=='Advogado Responsável pela Audiência':
                val=normalize_adv(prsv.get('adv'))
                if not val:
                    if resp and resp not in ('nan',''):
                        val=normalize_adv(resp)
            elif nome=='Dados dos Correspondentes':
                val=f'=IFERROR(INDEX(SEMANA!$G:$G,MATCH($D{er},SEMANA!$D:$D,0)),"")'
            elif nome=='Acompanhamento': val=prsv.get('acomp')
            elif nome=='Observações':   val=prsv.get('obs')
            elif nome=='_CANC':
                val=f'=IFERROR(INDEX(SEMANA!$P:$P,MATCH($D{er},SEMANA!$D:$D,0)),"")'
            else:
                src=SRC_MAP.get(nome,nome)
                if src in df_g.columns:
                    v=row[src]
                    if not(v is None or(isinstance(v,float) and pd.isna(v))): val=v
            c=ws_g.cell(er,ci,val)
            c.font=Font(name='Arial',size=9,color='000000',bold=False,italic=False)
            c.alignment=Alignment(vertical='center',wrap_text=(nome in WRAP_G))
            if nome.startswith('_'): c.border=Border(); c.font=Font(name='Arial',size=9,color='FFFFFF')
            else: c.border=lb()
            if nome=='Data/hora de início' and val is not None and not isinstance(val,str):
                c.number_format='DD/MM/YYYY HH:MM'
        ws_g.row_dimensions[er].height=18

    ws_g.column_dimensions['W'].hidden=True

    # CF GERAL
    N=len(df_g)+500; DR=f'A3:V{N+2}'
    ws_g.conditional_formatting.add(DR,OXLRule(type='expression',
        formula=[f'AND($A3<>"",$A3<{_S1})'],
        dxf=DifferentialStyle(fill=PatternFill('solid',start_color='EBEBEB'),
                              font=Font(name='Arial',size=9,color='999999')),priority=10))
    ws_g.conditional_formatting.add(DR,OXLRule(type='expression',
        formula=['ISNUMBER(SEARCH("cancelad",$W3))'],
        dxf=DifferentialStyle(fill=PatternFill('solid',start_color='FFDDE1'),
                              font=Font(name='Arial',size=9,color='000000')),priority=1))
    ws_g.conditional_formatting.add(DR,OXLRule(type='expression',
        formula=['ISNUMBER(SEARCH("CONTROLADORIA",$G3))'],
        dxf=DifferentialStyle(fill=PatternFill('solid',start_color=ORANGE_ROW),
                              font=Font(name='Arial',size=9,color='000000')),priority=2))
    ws_g.conditional_formatting.add(DR,OXLRule(type='expression',
        formula=['$I3="Sim"'],
        dxf=DifferentialStyle(fill=PatternFill('solid',start_color=GREEN_ROW),
                              font=Font(name='Arial',size=9,color='000000')),priority=3))
    # CF extra: divergência (nome não reconhecido)
    if divs:
        ws_g.conditional_formatting.add(DR,OXLRule(type='expression',
            formula=['ISNUMBER(SEARCH("⚠",GERAL!$G3))'],
            dxf=DifferentialStyle(fill=PatternFill('solid',start_color=DIV_RED),
                                  font=Font(name='Arial',size=9,color=DIV_DARK,bold=True)),priority=0))

    dv_acomp=DataValidation(type='list',formula1='"Sim,Não"',allow_blank=True,showDropDown=False)
    dv_acomp.sqref='I3:I5000'; ws_g.add_data_validation(dv_acomp)
    dv_adv=DataValidation(type='list',formula1=f'CONFIG!$A$2:$A${len(all_advs)+1}',allow_blank=True,showDropDown=False)
    dv_adv.sqref='G3:G5000'; ws_g.add_data_validation(dv_adv)
    dv_mod=DataValidation(type='list',formula1='"PRESENCIAL,VIRTUAL,HÍBRIDA"',allow_blank=True,showDropDown=False)
    dv_mod.sqref='M3:M5000'; ws_g.add_data_validation(dv_mod)
    ws_g.freeze_panes='A3'; ws_g.auto_filter.ref='A2:V5000'

    # ── ABA SEMANA ─────────────────────────────────────────────────────────────
    ws_sem=wb.create_sheet('SEMANA'); ws_sem.sheet_properties.tabColor='FF7F00'
    SEM_COLS=[
        ('Data/hora de início',16,'S'),('Natureza',13,'S'),('Descrição',44,'S'),
        ('Número de CNJ',24,'S'),('Tipo / Subtipo',20,'S'),('Contratação',22,'A'),
        ('Dados dos Correspondentes',44,'C'),('Preposto',30,'C'),('Observações',34,'A'),
        ('Cliente Processo',38,'S'),('Contrário principal',28,'S'),
        ('Modalidade',14,'A'),('Local',42,'A'),('Cidade',18,'S'),('UF',8,'S'),
        ('Cancelamento',18,'A'),
    ]
    SEM_CSTY={'S':(SISTEMA_BG,'1F3864'),'A':(ADVOG_BG,'375623'),'C':(CTRL_BG,'843C0C')}
    SEM_SRC={'Data/hora de início':'Data/hora de início','Natureza':'Natureza',
             'Descrição':'Descrição','Número de CNJ':'Número de CNJ',
             'Tipo / Subtipo':'Tipo / Subtipo','Cliente Processo':'Cliente Processo',
             'Contrário principal':'Contrário principal','Modalidade':'Modalidade',
             'Local':'Local','Cidade':'Cidade','UF':'UF'}
    ws_sem.row_dimensions[2].height=34
    for ci,(nome,larg,tipo) in enumerate(SEM_COLS,1):
        bg,fg=SEM_CSTY.get(tipo,(SISTEMA_BG,'1F3864'))
        c=ws_sem.cell(2,ci,nome)
        c.font=Font(name='Arial',bold=True,color=fg,size=9)
        c.fill=PatternFill('solid',start_color=bg)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=tb()
        ws_sem.column_dimensions[get_column_letter(ci)].width=larg
    EDIT_SEM={'Contratação','Dados dos Correspondentes','Preposto','Observações','Cancelamento'}
    for ri,row in df_g.iterrows():
        er=ri+3
        cnj=str(row.get('Número de CNJ','')).strip(); cli=str(row.get('Cliente Processo','')).strip()
        data=str(row.get('Data/hora de início','')).strip()[:10]; desc=str(row.get('Descrição','')).strip()[:60]
        prsv_s=(preserved.get((cnj,cli,data,desc)) or preserved_2.get((cnj,cli)) or preserved_1.get(cnj) or {})
        for ci,(nome,_,tipo) in enumerate(SEM_COLS,1):
            val=None
            if nome=='Dados dos Correspondentes': val=prsv_s.get('dados')
            elif nome=='Contratação': val=prsv_s.get('contrat')
            elif nome=='Preposto': val=prsv_s.get('preposto')
            elif nome=='Observações': val=prsv_s.get('obs')
            elif nome=='Cancelamento': val=None
            else:
                src=SEM_SRC.get(nome,nome)
                if src in df_g.columns:
                    v=row.get(src)
                    if not(v is None or(isinstance(v,float) and pd.isna(v))): val=v
            c=ws_sem.cell(er,ci,val)
            c.font=Font(name='Arial',size=9,color='000000',bold=False)
            c.alignment=Alignment(vertical='center',wrap_text=False if nome in('Dados dos Correspondentes','Preposto') else (nome in {'Descrição','Local','Observações'}))
            c.border=lb()
            if nome=='Data/hora de início' and val is not None and not isinstance(val,str):
                c.number_format='DD/MM/YYYY HH:MM'
        ws_sem.row_dimensions[er].height=30
    dv_cont=DataValidation(type='list',formula1='"Indicar adv,Indicar adv e preposto,Indicar preposto"',allow_blank=True,showDropDown=False)
    dv_cont.sqref='F3:F5000'; ws_sem.add_data_validation(dv_cont)
    dv_canc=DataValidation(type='list',formula1='"Cancelado,Manter,Redesignado"',allow_blank=True,showDropDown=False)
    dv_canc.sqref='P3:P5000'; ws_sem.add_data_validation(dv_canc)
    dv_mod_s=DataValidation(type='list',formula1='"PRESENCIAL,VIRTUAL,HÍBRIDA"',allow_blank=True,showDropDown=False)
    dv_mod_s.sqref='L3:L5000'; ws_sem.add_data_validation(dv_mod_s)
    ws_sem.freeze_panes='A3'; ws_sem.auto_filter.ref='A2:P5000'

    # ── DASHBOARD GERAL ────────────────────────────────────────────────────────
    ws_dg=wb.create_sheet('DASHBOARD GERAL'); ws_dg.sheet_properties.tabColor=NAVY
    ws_dg.sheet_view.showGridLines=False
    for ltr,w in [('A',2),('B',26),('C',11),('D',11),('E',11),('F',11),('G',11),('H',11),('I',11),('J',11),('K',11),('L',11),('M',11),('N',2)]:
        ws_dg.column_dimensions[ltr].width=w
    ws_dg.row_dimensions[1].height=46
    ws_dg.merge_cells('B1:M1')
    c=ws_dg['B1']; c.value='PAUTA DE AUDIÊNCIAS  ·  DASHBOARD GERAL'
    c.font=Font(name='Arial',bold=True,size=20,color=WHITE)
    c.fill=PatternFill('solid',start_color=NAVY)
    c.alignment=Alignment(horizontal='center',vertical='center')

    # Alerta de divergências no dashboard
    if divs:
        ws_dg.row_dimensions[2].height=28
        ws_dg.merge_cells('B2:M2')
        n_div=len(divs)
        c=ws_dg['B2']
        c.value=f'⚠️  ATENÇÃO — {n_div} DIVERGÊNCIA(S) DETECTADA(S) — Verifique a aba DIVERGÊNCIAS antes de distribuir o arquivo'
        c.font=Font(name='Arial',bold=True,size=10,color=DIV_DARK)
        c.fill=PatternFill('solid',start_color=DIV_RED)
        c.alignment=Alignment(horizontal='center',vertical='center')
        c.border=tb(DIV_DARK)
        r_start=3
    else:
        ws_dg.row_dimensions[2].height=10
        ws_dg.merge_cells('B2:M2')
        c=ws_dg['B2']; c.value='✅  Nenhuma divergência detectada'
        c.font=Font(name='Arial',size=9,color='145A32')
        c.fill=PatternFill('solid',start_color='D5F5E3')
        c.alignment=Alignment(horizontal='center',vertical='center')
        r_start=3

    ws_dg.row_dimensions[r_start].height=8
    for col in range(2,14): ws_dg.cell(r_start,col).fill=PatternFill('solid',start_color=MED_BLUE)
    r_start+=1

    # Label S1
    s1_label=f"SEMANA 1  ·  {s1_start.strftime('%d/%m')} a {(s1_start+datetime.timedelta(days=6)).strftime('%d/%m/%Y')}"
    ws_dg.row_dimensions[r_start].height=26
    ws_dg.merge_cells(f'B{r_start}:M{r_start}')
    c=ws_dg.cell(r_start,2,s1_label)
    c.font=Font(name='Arial',bold=True,size=13,color=WHITE)
    c.fill=PatternFill('solid',start_color=MED_BLUE)
    c.alignment=Alignment(horizontal='left',vertical='center'); c.border=tb()
    r_start+=1

    # 7 KPI cards S1
    KPI_S1=[
        ('TOTAL\nAUDIÊNCIAS',    f'=COUNTIFS({sc1()})','1F3864','D6EAF8',2,3),
        ('PRESENCIAIS', f'=COUNTIFS({sc1(f",GERAL!$M:$M,{_P}")})','1E8449','D5F5E3',4,5),
        ('VIRTUAIS',    f'=COUNTIFS({sc1(f",GERAL!$M:$M,{_V}")})','1A5276','D6EAF8',6,7),
        ('HÍBRIDAS',    f'=COUNTIFS({sc1(f",GERAL!$M:$M,{_H}")})','6E2C00','FDEBD0',8,9),
        ('CORRESPONDENTE\nNECESSÁRIO',f'=COUNTIFS({sc1(f",GERAL!$G:$G,{_CJ}")})',
         '7D4300','FFF3CD',10,11),
        ('SUPORTE', f'=COUNTIFS({sc1(f",GERAL!$G:$G,{_Q}SUPORTE{_Q}")})', '4A235A','F5D6FF',12,12),
        ('COM\nACOMPANHAMENTO',f'=COUNTIFS({sc1(f",GERAL!$I:$I,{_SIM}")})','145A32','D5F5E3',13,13),
    ]
    ws_dg.row_dimensions[r_start].height=22
    ws_dg.row_dimensions[r_start+1].height=50
    ws_dg.row_dimensions[r_start+2].height=8
    for label,fml,dk,lk,cs,ce in KPI_S1:
        ws_dg.merge_cells(start_row=r_start,start_column=cs,end_row=r_start,end_column=ce)
        c=ws_dg.cell(r_start,cs,label)
        c.font=Font(name='Arial',bold=True,size=8,color=dk)
        c.fill=PatternFill('solid',start_color=lk)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=tb()
        ws_dg.merge_cells(start_row=r_start+1,start_column=cs,end_row=r_start+1,end_column=ce)
        c=ws_dg.cell(r_start+1,cs,fml)
        c.font=Font(name='Arial',bold=True,size=28,color=dk)
        c.fill=PatternFill('solid',start_color=lk)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
        ws_dg.merge_cells(start_row=r_start+2,start_column=cs,end_row=r_start+2,end_column=ce)
        ws_dg.cell(r_start+2,cs).fill=PatternFill('solid',start_color=dk)
    r_start+=3

    # Tabela por coordenador S1
    r=r_start
    ws_dg.row_dimensions[r].height=22
    ws_dg.merge_cells(f'B{r}:M{r}')
    c=ws_dg.cell(r,2,'  SEMANA 1  ·  AUDIÊNCIAS POR COORDENADOR')
    c.font=Font(name='Arial',bold=True,size=11,color=WHITE)
    c.fill=PatternFill('solid',start_color=NAVY); c.alignment=Alignment(horizontal='left',vertical='center'); c.border=tb()
    r+=1
    ws_dg.row_dimensions[r].height=28
    for cs,label,span in [(2,'COORDENADOR',5),(7,'TOTAL',1),(8,'PRES.',1),(9,'VIRT.',1),(10,'HÍBR.',1),(11,'CTRL.',1),(12,'SUPORTE',1),(13,'C/ACOMP.',1)]:
        ws_dg.merge_cells(start_row=r,start_column=cs,end_row=r,end_column=cs+span-1)
        c=ws_dg.cell(r,cs,label)
        c.font=Font(name='Arial',bold=True,size=9,color=WHITE)
        c.fill=PatternFill('solid',start_color=MED_BLUE)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=tb()
    r+=1; coord_s1_start=r

    for idx,(coord,info) in enumerate(COORDS.items()):
        if coord=='CONTROLADORIA JURÍDICA': continue
        cf=f'*{coord.split()[0]}*' if any(ord(ch)>127 for ch in coord) else coord
        alt=GRAY_ALT if idx%2==0 else WHITE
        ws_dg.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
        c=ws_dg.cell(r,2,coord)
        c.font=Font(name='Arial',size=9); c.fill=PatternFill('solid',start_color=alt)
        c.alignment=Alignment(horizontal='left',vertical='center'); c.border=tb()
        for col,cond in [
            (7,''),(8,f',GERAL!$M:$M,{_P}'),(9,f',GERAL!$M:$M,{_V}'),(10,f',GERAL!$M:$M,{_H}'),
            (11,f',GERAL!$G:$G,{_CJ}'),(12,',GERAL!$G:$G,"SUPORTE"'),(13,f',GERAL!$I:$I,{_SIM}'),
        ]:
            c=ws_dg.cell(r,col,f'=COUNTIFS({sc1(f",GERAL!$F:$F,{_Q}{cf}{_Q}{cond}")})')
            c.font=Font(name='Arial',size=9); c.fill=PatternFill('solid',start_color=alt)
            c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
        ws_dg.row_dimensions[r].height=18; r+=1

    # Total
    ws_dg.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    c=ws_dg.cell(r,2,'TOTAL SEMANA 1')
    c.font=Font(name='Arial',bold=True,size=9,color=WHITE)
    c.fill=PatternFill('solid',start_color=NAVY)
    c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
    for col in [7,8,9,10,11,12,13]:
        cl=get_column_letter(col)
        c=ws_dg.cell(r,col,f'=SUM({cl}{coord_s1_start}:{cl}{r-1})')
        c.font=Font(name='Arial',bold=True,size=9,color=WHITE)
        c.fill=PatternFill('solid',start_color=NAVY)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
    ws_dg.row_dimensions[r].height=20; r+=2

    # Prévia S2
    s2_label=f"PRÉVIA — SEMANA 2  ·  {(s1_start+datetime.timedelta(days=7)).strftime('%d/%m')} a {(s1_start+datetime.timedelta(days=13)).strftime('%d/%m/%Y')}"
    ws_dg.row_dimensions[r].height=6
    for col in range(2,14): ws_dg.cell(r,col).fill=PatternFill('solid',start_color=S2_DARK)
    r+=1
    ws_dg.row_dimensions[r].height=22
    ws_dg.merge_cells(f'B{r}:M{r}')
    c=ws_dg.cell(r,2,s2_label)
    c.font=Font(name='Arial',bold=True,size=11,color=WHITE)
    c.fill=PatternFill('solid',start_color=S2_DARK)
    c.alignment=Alignment(horizontal='left',vertical='center'); c.border=tb()
    r+=1
    S2_KPI=[
        ('TOTAL PRÉVIA',f'=COUNTIFS({sc2()})','555555'),
        ('PRESENCIAL',  f'=COUNTIFS({sc2(f",GERAL!$M:$M,{_P}")})','555555'),
        ('VIRTUAL',     f'=COUNTIFS({sc2(f",GERAL!$M:$M,{_V}")})','555555'),
        ('HÍBRIDA',     f'=COUNTIFS({sc2(f",GERAL!$M:$M,{_H}")})','555555'),
        ('CORRESP.',    f'=COUNTIFS({sc2(f",GERAL!$G:$G,{_CJ}")})','555555'),
        ('SUPORTE', f'=COUNTIFS({sc2(f",GERAL!$G:$G,{_Q}SUPORTE{_Q}")})', '555555'),
        ('C/ACOMP.',    f'=COUNTIFS({sc2(f",GERAL!$I:$I,{_SIM}")})','555555'),
    ]
    ws_dg.row_dimensions[r].height=16
    for i,(label,fml,dk) in enumerate(S2_KPI):
        cs=2+i*2 if i<6 else 13
        ce=cs+1 if i<5 else cs
        ws_dg.merge_cells(start_row=r,start_column=cs,end_row=r,end_column=ce)
        c=ws_dg.cell(r,cs,label)
        c.font=Font(name='Arial',bold=True,size=8,color=S2_MED)
        c.fill=PatternFill('solid',start_color=S2_ACCENT)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
    r+=1
    ws_dg.row_dimensions[r].height=34
    for i,(label,fml,dk) in enumerate(S2_KPI):
        cs=2+i*2 if i<6 else 13
        ce=cs+1 if i<5 else cs
        ws_dg.merge_cells(start_row=r,start_column=cs,end_row=r,end_column=ce)
        c=ws_dg.cell(r,cs,fml)
        c.font=Font(name='Arial',bold=True,size=20,color=S2_DARK)
        c.fill=PatternFill('solid',start_color=S2_LIGHT)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
    r+=2

    ws_dg.freeze_panes='B4'

    # ── DASH. COORD. ───────────────────────────────────────────────────────────
    ws_dc=wb.create_sheet('DASH. COORD.'); ws_dc.sheet_properties.tabColor=MED_BLUE
    ws_dc.sheet_view.showGridLines=False
    for ltr,w in [('A',2),('B',44),('C',12),('D',12),('E',12),('F',12),('G',3),('H',3),('I',12),('J',12),('K',12),('L',2)]:
        ws_dc.column_dimensions[ltr].width=w
    ws_dc.row_dimensions[1].height=46
    ws_dc.merge_cells('B1:K1')
    c=ws_dc['B1']; c.value='AUDIÊNCIAS POR COORDENADOR E ADVOGADO'
    c.font=Font(name='Arial',bold=True,size=18,color=WHITE)
    c.fill=PatternFill('solid',start_color=NAVY)
    c.alignment=Alignment(horizontal='center',vertical='center')

    # Grupos dinâmicos
    _dyn=defaultdict(set)
    for ri,row in df_g.iterrows():
        coord_v=str(row.get('Coordenador','')).strip()
        adv_v=normalize_adv(str(row.get('Responsável pela Pasta','')).strip())
        if not adv_v:
            adv_v=_resolved.get(ri,{}).get('adv','')
        if coord_v and adv_v: _dyn[coord_v].add(adv_v)

    rdc=3
    ws_dc.row_dimensions[rdc].height=20
    ws_dc.merge_cells(f'B{rdc}:G{rdc}')
    c=ws_dc.cell(rdc,2,'⚡ SEMANA 1')
    c.font=Font(name='Arial',bold=True,size=10,color=WHITE); c.fill=PatternFill('solid',start_color=MED_BLUE)
    c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
    ws_dc.cell(rdc,8).fill=PatternFill('solid',start_color=S2_ACCENT)
    ws_dc.merge_cells(f'I{rdc}:K{rdc}')
    c=ws_dc.cell(rdc,9,'🗓 PRÉVIA — SEMANA 2')
    c.font=Font(name='Arial',bold=True,size=10,color=WHITE); c.fill=PatternFill('solid',start_color=S2_DARK)
    c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
    rdc+=1

    for coord,info in COORDS.items():
        if coord=='CONTROLADORIA JURÍDICA': continue
        active=sorted(_dyn.get(coord,set()))
        if not active: continue
        dk,lk=info['dark'],info['light']
        cf=f'*{coord.split()[0]}*' if any(ord(ch)>127 for ch in coord) else coord
        ws_dc.row_dimensions[rdc].height=24
        ws_dc.merge_cells(f'B{rdc}:K{rdc}')
        c=ws_dc.cell(rdc,2,f'  {coord}')
        c.font=Font(name='Arial',bold=True,size=11,color=WHITE)
        c.fill=PatternFill('solid',start_color=dk)
        c.alignment=Alignment(horizontal='left',vertical='center'); c.border=tb(); rdc+=1

        ws_dc.row_dimensions[rdc].height=24
        for cs,label in [(2,'ADVOGADO'),(3,'TOTAL'),(4,'PRES.'),(5,'VIRT.'),(6,'ACOMP.')]:
            c=ws_dc.cell(rdc,cs,label)
            c.font=Font(name='Arial',bold=True,size=8,color=dk)
            c.fill=PatternFill('solid',start_color=lk)
            c.alignment=Alignment(horizontal='center' if cs>2 else 'left',vertical='center',wrap_text=True); c.border=tb()
        ws_dc.cell(rdc,8).fill=PatternFill('solid',start_color=S2_ACCENT); ws_dc.cell(rdc,8).border=tb()
        for cs,label in [(9,'TOTAL'),(10,'PRES.'),(11,'VIRT.')]:
            c=ws_dc.cell(rdc,cs,label)
            c.font=Font(name='Arial',bold=True,size=8,color=S2_DARK)
            c.fill=PatternFill('solid',start_color=S2_ACCENT)
            c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=tb()
        rdc+=1; adv_start=rdc

        for i,adv in enumerate(active):
            alt=GRAY_ALT if i%2==0 else WHITE
            c=ws_dc.cell(rdc,2,adv)
            c.font=Font(name='Arial',size=9); c.fill=PatternFill('solid',start_color=alt)
            c.alignment=Alignment(horizontal='left',vertical='center',indent=1); c.border=tb()
            for col,cond in [(3,''),(4,f',GERAL!$M:$M,{_P}'),(5,f',GERAL!$M:$M,{_V}'),(6,f',GERAL!$I:$I,{_SIM}')]:
                fml=f'=COUNTIFS({sc1(f",GERAL!$G:$G,{_Q}{adv}{_Q}{cond}")})'
                c=ws_dc.cell(rdc,col,fml)
                c.font=Font(name='Arial',size=9); c.fill=PatternFill('solid',start_color=alt)
                c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
            ws_dc.cell(rdc,8).fill=PatternFill('solid',start_color=S2_ACCENT); ws_dc.cell(rdc,8).border=lb()
            for col,cond in [(9,''),(10,f',GERAL!$M:$M,{_P}'),(11,f',GERAL!$M:$M,{_V}')]:
                fml=f'=COUNTIFS({sc2(f",GERAL!$F:$F,{_Q}{cf}{_Q}{cond}")})'
                c=ws_dc.cell(rdc,col,'')
                c.font=Font(name='Arial',size=9,color=S2_DARK)
                c.fill=PatternFill('solid',start_color=S2_LIGHT if i%2==0 else S2_ACCENT)
                c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
            ws_dc.row_dimensions[rdc].height=18; rdc+=1

        ws_dc.row_dimensions[rdc].height=18
        c=ws_dc.cell(rdc,2,'SUBTOTAL')
        c.font=Font(name='Arial',bold=True,size=9,color=WHITE)
        c.fill=PatternFill('solid',start_color=dk)
        c.alignment=Alignment(horizontal='left',vertical='center'); c.border=tb()
        for col in [3,4,5,6]:
            cl=get_column_letter(col)
            c=ws_dc.cell(rdc,col,f'=SUM({cl}{adv_start}:{cl}{rdc-1})')
            c.font=Font(name='Arial',bold=True,size=9,color=WHITE)
            c.fill=PatternFill('solid',start_color=dk)
            c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
        ws_dc.cell(rdc,8).fill=PatternFill('solid',start_color=S2_DARK); ws_dc.cell(rdc,8).border=tb()
        for col,cond in [(9,''),(10,f',GERAL!$M:$M,{_P}'),(11,f',GERAL!$M:$M,{_V}')]:
            c=ws_dc.cell(rdc,col,f'=COUNTIFS({sc2(f",GERAL!$F:$F,{_Q}{cf}{_Q}{cond}")})')
            c.font=Font(name='Arial',bold=True,size=9,color=WHITE)
            c.fill=PatternFill('solid',start_color=S2_DARK)
            c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
        rdc+=2

    ws_dc.freeze_panes='B5'

    # ── DASH. CJ ───────────────────────────────────────────────────────────────
    ws_ctrl=wb.create_sheet('DASH. CJ'); ws_ctrl.sheet_properties.tabColor=RED_DARK
    ws_ctrl.sheet_view.showGridLines=False
    for ltr,w in [('A',2),('B',17),('C',40),('D',28),('E',14),('F',42),('G',42),('H',30),('I',2)]:
        ws_ctrl.column_dimensions[ltr].width=w
    ws_ctrl.row_dimensions[1].height=42
    ws_ctrl.merge_cells('B1:H1')
    c=ws_ctrl['B1']; c.value='CONTROLADORIA JURÍDICA  ·  CORRESPONDENTES E PREPOSTOS'
    c.font=Font(name='Arial',bold=True,size=15,color=WHITE)
    c.fill=PatternFill('solid',start_color=RED_DARK)
    c.alignment=Alignment(horizontal='center',vertical='center')

    S1_KPIS=[
        ('TOTAL COM\nCORRESPONDENTE',f'=COUNTIFS({sc1(f",GERAL!$G:$G,{_CJ}")})',
         '7D4300','FFF3CD',2,4),
        ('JÁ\nATRIBUÍDOS',
         f'=SUMPRODUCT((GERAL!$A$3:$A$1000>={_S1})*(GERAL!$A$3:$A$1000<{_S2})*(GERAL!$G$3:$G$1000=CONFIG!$B$6)*(LEN(SEMANA!$G$3:$G$1000)>0))',
         '145A32','D5F5E3',5,6),
        ('AGUARDANDO\nATRIBUIÇÃO',
         f'=SUMPRODUCT((GERAL!$A$3:$A$1000>={_S1})*(GERAL!$A$3:$A$1000<{_S2})*(GERAL!$G$3:$G$1000=CONFIG!$B$6)*(LEN(SEMANA!$G$3:$G$1000)=0))',
         '922B21','FADBD8',7,8),
    ]
    ws_ctrl.row_dimensions[2].height=22
    ws_ctrl.merge_cells('B2:H2')
    c=ws_ctrl['B2']; c.value=f'⚡  SEMANA 1  —  {s1_start.strftime("%d/%m")} a {(s1_start+datetime.timedelta(days=6)).strftime("%d/%m/%Y")}'
    c.font=Font(name='Arial',bold=True,size=11,color=WHITE); c.fill=PatternFill('solid',start_color='8B1A1A'); c.alignment=Alignment(horizontal='left',vertical='center'); c.border=tb()
    ws_ctrl.row_dimensions[3].height=18; ws_ctrl.row_dimensions[4].height=40; ws_ctrl.row_dimensions[5].height=8
    for label,fml,dk,lk,cs,ce in S1_KPIS:
        ws_ctrl.merge_cells(start_row=3,start_column=cs,end_row=3,end_column=ce)
        c=ws_ctrl.cell(3,cs,label); c.font=Font(name='Arial',bold=True,size=9,color=dk)
        c.fill=PatternFill('solid',start_color=lk); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=tb()
        ws_ctrl.merge_cells(start_row=4,start_column=cs,end_row=4,end_column=ce)
        c=ws_ctrl.cell(4,cs,fml); c.font=Font(name='Arial',bold=True,size=26,color=dk)
        c.fill=PatternFill('solid',start_color=lk); c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tb()
        ws_ctrl.merge_cells(start_row=5,start_column=cs,end_row=5,end_column=ce)
        ws_ctrl.cell(5,cs).fill=PatternFill('solid',start_color=dk)

    # Lista CTRL
    r=7
    ws_ctrl.row_dimensions[r].height=24
    ws_ctrl.merge_cells(f'B{r}:H{r}')
    c=ws_ctrl.cell(r,2,'  LISTA — SEMANA 1  ·  REQUEREM CONTRATAÇÃO DE CORRESPONDENTE / PREPOSTO')
    c.font=Font(name='Arial',bold=True,size=11,color=WHITE); c.fill=PatternFill('solid',start_color=RED_DARK)
    c.alignment=Alignment(horizontal='left',vertical='center'); c.border=tb(); r+=1
    ws_ctrl.row_dimensions[r].height=20
    for cp,label in [(2,'DATA'),(3,'CLIENTE'),(4,'CONTRÁRIO'),(5,'MODALIDADE'),(6,'LOCAL'),(7,'DADOS DO CORRESPONDENTE'),(8,'COORDENADOR')]:
        c=ws_ctrl.cell(r,cp,label); c.font=Font(name='Arial',bold=True,size=9,color=WHITE)
        c.fill=PatternFill('solid',start_color=NAVY); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=tb()
    r+=1

    _ctrl_rows=[]
    for ri,row in df_g.iterrows():
        cnj_c=str(row.get('Número de CNJ','')).strip(); cli_c=str(row.get('Cliente Processo','')).strip()
        data_c=str(row.get('Data/hora de início','')).strip()[:10]; desc_c=str(row.get('Descrição','')).strip()[:60]
        prsv_c=(preserved.get((cnj_c,cli_c,data_c,desc_c)) or preserved_2.get((cnj_c,cli_c)) or preserved_1.get(cnj_c) or {})
        adv_c=normalize_adv(_resolved.get(ri,{}).get('adv') or prsv_c.get('adv') or row.get('Advogado Responsável pela Audiência'))
        if adv_c and 'CONTROLADORIA' in str(adv_c).upper():
            row_copy=row.copy()
            row_copy['Advogado Responsável pela Audiência']=adv_c
            dados_c=_resolved.get(ri,{}).get('dados') or prsv_c.get('dados')
            if dados_c: row_copy['Dados dos Correspondentes']=dados_c
            _ctrl_rows.append(row_copy)

    for ri2,row in pd.DataFrame(_ctrl_rows).sort_values('Data/hora de início').iterrows() if _ctrl_rows else []:
        alt=GRAY_ALT if ri2%2==0 else WHITE
        dv=row.get('Data/hora de início')
        c=ws_ctrl.cell(r,2,dv if not(dv is None or(isinstance(dv,float) and pd.isna(dv))) else None)
        c.font=Font(name='Arial',size=9); c.fill=PatternFill('solid',start_color=alt)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=lb()
        if dv is not None: c.number_format='DD/MM/YYYY'
        for cp,col in [(3,'Cliente Processo'),(4,'Contrário principal'),(5,'Modalidade'),(6,'Local'),(7,'Dados dos Correspondentes'),(8,'Coordenador')]:
            v=row.get(col); val=None if(v is None or(isinstance(v,float) and pd.isna(v))) else v
            c=ws_ctrl.cell(r,cp,val); c.font=Font(name='Arial',size=9); c.fill=PatternFill('solid',start_color=alt)
            c.alignment=Alignment(vertical='center',wrap_text=(cp in[6,7])); c.border=lb()
        ws_ctrl.row_dimensions[r].height=20; r+=1
    ws_ctrl.freeze_panes='B7'

    # ── ABA DIVERGÊNCIAS ───────────────────────────────────────────────────────
    ws_div=wb.create_sheet('DIVERGÊNCIAS'); ws_div.sheet_properties.tabColor=DIV_DARK if divs else '145A32'
    ws_div.sheet_view.showGridLines=True
    for ltr,w in [('A',18),('B',22),('C',35),('D',35),('E',14),('F',50),('G',8)]:
        ws_div.column_dimensions[ltr].width=w
    ws_div.row_dimensions[1].height=40
    ws_div.merge_cells('A1:G1')
    c=ws_div['A1']
    if divs:
        c.value=f'⚠️  DIVERGÊNCIAS DETECTADAS — {len(divs)} ocorrência(s)  ·  Gerado em {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}'
        c.fill=PatternFill('solid',start_color=DIV_RED)
        c.font=Font(name='Arial',bold=True,size=12,color=DIV_DARK)
    else:
        c.value=f'✅  Nenhuma divergência detectada  ·  Gerado em {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}'
        c.fill=PatternFill('solid',start_color='D5F5E3')
        c.font=Font(name='Arial',bold=True,size=12,color='145A32')
    c.alignment=Alignment(horizontal='center',vertical='center')

    ws_div.row_dimensions[2].height=24
    for ci,(lbl,w) in enumerate(zip(['TIPO','CAMPO','VALOR ENCONTRADO','VALOR SUGERIDO','GRAVIDADE','DESCRIÇÃO','QTD'],
                                     [18,22,35,35,14,50,8]),1):
        c=ws_div.cell(2,ci,lbl)
        c.font=Font(name='Arial',bold=True,size=9,color=WHITE)
        c.fill=PatternFill('solid',start_color=NAVY if not divs else DIV_DARK)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=tb()

    for ri,d in enumerate(divs,3):
        bg=DIV_RED if d['gravidade']=='ALTA' else ('FFF3CD' if d['gravidade']=='MÉDIA' else 'FFFDE7')
        for ci,val in enumerate([d['tipo'],d['campo'],d['valor'],d.get('valor_sugerido',''),
                                   d['gravidade'],d['descricao'],d['ocorrencias']],1):
            c=ws_div.cell(ri,ci,val)
            c.font=Font(name='Arial',size=9,bold=(d['gravidade']=='ALTA'))
            c.fill=PatternFill('solid',start_color=bg)
            c.alignment=Alignment(vertical='center',wrap_text=True); c.border=lb()
        ws_div.row_dimensions[ri].height=20

    # ── CANCELADAS ─────────────────────────────────────────────────────────────
    ws_canc=wb.create_sheet('CANCELADAS'); ws_canc.sheet_properties.tabColor=RED_DARK
    CANC_COLS=[('Data/hora de início',16),('Natureza',13),('Descrição',42),('Número de CNJ',24),
               ('Tipo / Subtipo',20),('Advogado Responsável pela Audiência',30),
               ('Observações Complementares',38),('Cliente Processo',36),('Contrário principal',28),
               ('Modalidade',14),('Local',40),('Classificação do Processo',20),
               ('Responsável pela Pasta',28),('Ação',22),('Órgão',38),('Cidade',16),('UF',8),('Status',12)]
    ws_canc.row_dimensions[1].height=34
    for ci,(nome,larg) in enumerate(CANC_COLS,1):
        hcell(ws_canc,1,ci,nome,fg=WHITE,bg=RED_DARK,size=9)
        ws_canc.column_dimensions[get_column_letter(ci)].width=larg
    for ri,row in df_c.iterrows():
        er=ri+2
        for ci,(nome,_) in enumerate(CANC_COLS,1):
            val=None
            if nome in df_c.columns:
                v=row[nome]; val=None if(v is None or(isinstance(v,float) and pd.isna(v))) else v
            c=ws_canc.cell(er,ci,val); c.font=Font(name='Arial',size=9); c.alignment=Alignment(vertical='center'); c.border=lb()
            if nome=='Data/hora de início' and val is not None and not isinstance(val,str):
                c.number_format='DD/MM/YYYY HH:MM'
        ws_canc.row_dimensions[er].height=18
    ws_canc.freeze_panes='A2'; ws_canc.auto_filter.ref=f'A1:{get_column_letter(len(CANC_COLS))}1'

    # ── VIAGENS. DILIG. E SUPORTE ──────────────────────────────────────────────
    ws_vds=wb.create_sheet('VIAGENS. DILIG. E SUPORTE'); ws_vds.sheet_properties.tabColor='2E75B6'
    VDS_COLS=[('Advogado',30),('Viagens. Diligências e Suporte',28),('Local / Disponibilidade',36),('Data / Período',22),('Natureza / Atividade',22),('Suporte ADM · Observações',38)]
    ws_vds.row_dimensions[1].height=36
    for ci,(nome,larg) in enumerate(VDS_COLS,1):
        hcell(ws_vds,1,ci,nome,fg=WHITE,bg='2E75B6',size=9)
        ws_vds.column_dimensions[get_column_letter(ci)].width=larg
    ws_vds.freeze_panes='A2'

    # ── ORDEM DAS ABAS ─────────────────────────────────────────────────────────
    ORDER=['DASHBOARD GERAL','GERAL','DASH. COORD.','DASH. CJ','SEMANA','CANCELADAS',
           'DIVERGÊNCIAS','VIAGENS. DILIG. E SUPORTE','CONFIG']
    ordered=[wb[n] for n in ORDER if n in wb.sheetnames]
    remaining=[s for s in wb._sheets if s not in ordered]
    wb._sheets=ordered+remaining

    # Salvar
    buf=io.BytesIO(); wb.save(buf); output_bytes=buf.getvalue()

    # Resumo
    resumo={
        'total_pendentes':len(df_g),
        'total_canceladas':len(df_c),
        's1_start':s1_start.strftime('%d/%m/%Y'),
        's1_end':(s1_start+datetime.timedelta(days=6)).strftime('%d/%m/%Y'),
        's2_start':(s1_start+datetime.timedelta(days=7)).strftime('%d/%m/%Y'),
        's2_end':(s1_start+datetime.timedelta(days=13)).strftime('%d/%m/%Y'),
        'ctrl_pendentes':len(_ctrl_rows),
        'preservados_chave':len(preserved),
        'preservados_cnj_cli':len(preserved_2),
        'divergencias':len(divs),
        'gerado_em':datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
    }

    return output_bytes, resumo, divs
